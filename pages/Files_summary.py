import streamlit as st
import os
import PyPDF2
from openai import AzureOpenAI
from dotenv import load_dotenv
import pandas as pd
import docx
import openpyxl

# Specify the path to your .env file
env_path = os.path.join(os.path.dirname(__file__), '..', 'config', '.env')

# Load environment variables from the specified path
load_dotenv(dotenv_path=env_path)

# Set up Azure OpenAI client
client = AzureOpenAI(
    api_key=os.getenv("AZURE_OPENAI_KEY"),  
    api_version="2023-05-15",
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT")
)

def limit_tokens(text, max_chars=120000):
    return text[:max_chars]

def get_pdf_text(pdf_path):
    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
            if len(text) > 120000:
                break
    return limit_tokens(text)

def get_docx_text(docx_path):
    doc = docx.Document(docx_path)
    text = ""
    for para in doc.paragraphs:
        text += para.text + "\n"
        if len(text) > 120000:
            break
    return limit_tokens(text)

def get_excel_text(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join(str(cell) for cell in row if cell is not None) + "\n"
            if len(text) > 120000:
                break
        if len(text) > 120000:
            break
    return limit_tokens(text)

def summarize_text(text, max_tokens=150):
    response = client.chat.completions.create(
        model=os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME"),
        messages=[
            {"role": "system", "content": "You are a helpful assistant that summarizes text."},
            {"role": "user", "content": f"Please summarize the following text in about {max_tokens} tokens:\n\n{text}"}
        ],
        max_tokens=max_tokens,
        temperature=0.5,
    )
    return response.choices[0].message.content.strip()

def get_supported_files(folder_path):
    supported_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(('.pdf', '.docx', '.xlsx', '.xls')):
                supported_files.append(os.path.join(root, file))
    return supported_files

def main():
    st.title("Files Summarizer using TBH-Azure OpenAI")
    
    folder_path = st.text_input("Enter the path to the folder:")
    
    if folder_path and os.path.isdir(folder_path):
        supported_files = get_supported_files(folder_path)
        
        if supported_files:
            st.write(f"Found {len(supported_files)} supported files in the folder and its subfolders.")
            
            summaries = []
            
            progress_bar = st.progress(0)
            for i, file_path in enumerate(supported_files):
                try:
                    if file_path.lower().endswith('.pdf'):
                        text = get_pdf_text(file_path)
                    elif file_path.lower().endswith('.docx'):
                        text = get_docx_text(file_path)
                    elif file_path.lower().endswith(('.xlsx', '.xls')):
                        text = get_excel_text(file_path)
                    else:
                        continue

                    summary = summarize_text(text)
                    relative_path = os.path.relpath(file_path, folder_path)
                    summaries.append({
                        "File Name": os.path.basename(file_path),
                        "Relative Path": relative_path,
                        "Summary": summary
                    })
                except Exception as e:
                    st.error(f"Error processing {file_path}: {str(e)}")
                progress_bar.progress((i + 1) / len(supported_files))
            
            # Create a DataFrame and display it
            df = pd.DataFrame(summaries)
            st.dataframe(df)
            
            # Option to download the summary as a CSV file
            csv = df.to_csv(index=False)
            st.download_button(
                label="Download summaries as CSV",
                data=csv,
                file_name="file_summaries.csv",
                mime="text/csv",
            )
        else:
            st.warning("No supported files found in the specified folder or its subfolders.")
    elif folder_path:
        st.error("Invalid folder path. Please enter a valid path.")

if __name__ == "__main__":
    main()